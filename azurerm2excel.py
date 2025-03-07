import datetime
import json
import os
import re
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def parse_attributes(attributes, parent_key=""):
    attribute_list = []
    for key, value in attributes.items():
        if isinstance(value, dict):
            attribute_list.extend(
                parse_attributes(value, f"{parent_key}{key}.")
            )  # Recursive call for nested dictionaries
        elif isinstance(value, list):
            for i, item in enumerate(value):
                if isinstance(item, dict):
                    attribute_list.extend(
                        parse_attributes(item, f"{parent_key}{key}[{i}].")
                    )  # Recursive call for nested dictionaries in lists
                else:
                    attribute_list.append(
                        (f"{parent_key}{key}[{i}]", item)
                    )  # Add list item as a separate attribute
        else:
            attribute_list.append(
                (f"{parent_key}{key}", value)
            )  # Add simple key-value pair
    return attribute_list


def load_descriptions(
    description_folders,
):  # description_folder を description_folders に変更
    descriptions = {}
    for description_folder in description_folders:  # 各フォルダをループ
        for filename in os.listdir(description_folder):
            if filename.endswith(".json"):
                type_name = filename.split(".")[0]
                with open(
                    os.path.join(description_folder, filename), "r", encoding="utf-8"
                ) as f:
                    descriptions[type_name] = json.load(f)
    return descriptions


def apply_styles(
    ws, header_fill, header_font, default_font, left_alignment, thin_border
):
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in ws.iter_rows(min_row=2):  # Skip header row
        for cell in row:
            cell.font = default_font
            cell.alignment = left_alignment
            cell.border = thin_border
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    lines = str(cell.value).splitlines()
                    cell_max = max(len(line) for line in lines) if lines else 0
                    if cell_max > max_length:
                        max_length = cell_max
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width


def write_to_excel(resources_by_type, descriptions, output_folder):
    for resource_type, resources in resources_by_type.items():
        wb = Workbook()
        for resource in resources:
            # Limit sheet title to 30 characters
            sheet_title = resource["name"][:30]
            ws = wb.create_sheet(title=sheet_title)
            attribute_list = []
            for instance in resource["instances"]:
                attribute_list.extend(parse_attributes(instance["attributes"]))

            # Add header row
            header = ["Arguments", "Value", "Description"]
            ws.append(header)

            # Apply background color, font color, and font to the header row
            header_fill = PatternFill(
                start_color="0000FF", end_color="0000FF", fill_type="solid"
            )  # Blue background
            header_font = Font(
                color="FFFFFF", name="Yu Gothic UI", bold=True
            )  # White text, Yu Gothic UI font, bold
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            if resource_type == "azurerm_network_security_group":
                security_rules = [
                    attribute
                    for attribute in attribute_list
                    if "security_rule" in attribute[0]
                ]
                other_attributes = [
                    attribute
                    for attribute in attribute_list
                    # if "security_rule" not in attribute[0]
                ]

                # security_ruleだけ別のシートに出力
                if security_rules:
                    ws_security_rule = wb.create_sheet(title=f"{sheet_title}_rule")
                    # Add header row for security_rule sheet
                    ws_security_rule.append(
                        [
                            "rule_index",
                            "direction",
                            "priority",
                            "access",
                            "name",
                            "description",
                            "destination_address_prefix",
                            "destination_port_range",
                            "destination_port_ranges",
                            "protocol",
                            "source_address_prefix",
                            "source_port_range",
                        ]
                    )

                    rule_dict = defaultdict(dict)
                    for rule in security_rules:
                        rule_key = re.sub(r"\[\d+\]", "", rule[0])
                        rule_index = re.search(r"\[(\d+)\]", rule[0]).group(1)
                        if rule_key in rule_dict[rule_index]:
                            rule_dict[rule_index][rule_key] = str(rule_dict[rule_index][rule_key]) + f"\n{rule[1]}"
                        else:
                            rule_dict[rule_index][rule_key] = rule[1]

                    # Sort rules by priority
                    sorted_rules = sorted(
                        rule_dict.items(),
                        key=lambda x: (
                            x[1].get("security_rule.direction", ""),
                            int(x[1].get("security_rule.priority", 0)),
                        ),
                    )

                    for rule_index, rule_attrs in sorted_rules:
                        row = [f"security_rule{rule_index}"]
                        for key in [
                            "security_rule.direction",
                            "security_rule.priority",
                            "security_rule.access",
                            "security_rule.name",
                            "security_rule.description",
                            "security_rule.destination_address_prefix",
                            "security_rule.destination_port_range",
                            "security_rule.destination_port_ranges",
                            "security_rule.protocol",
                            "security_rule.source_address_prefix",
                            "security_rule.source_port_range",
                        ]:
                            row.append(str(rule_attrs.get(key, "")))
                        ws_security_rule.append(row)

                    # Apply styles to the new sheet
                    apply_styles(
                        ws_security_rule,
                        header_fill,
                        header_font,
                        default_font,
                        left_alignment,
                        thin_border,
                    )

                # Add other attributes to the original sheet
                for attribute in other_attributes:
                    attribute_path = attribute[0]
                    normalized_path = re.sub(r"\[\d+\]", "", attribute_path)
                    description = descriptions.get(resource_type, {}).get(
                        normalized_path, ""
                    )
                    value = str(attribute[1])
                    ws.append([attribute[0], value, description])
            elif resource_type == "azurerm_firewall_policy_rule_collection_group":
                application_rule_collections = [
                    attribute
                    for attribute in attribute_list
                    if "application_rule_collection" in attribute[0]
                ]
                network_rule_collections = [
                    attribute
                    for attribute in attribute_list
                    if "network_rule_collection" in attribute[0]
                ]
                nat_rule_collections = [
                    attribute
                    for attribute in attribute_list
                    if "nat_rule_collection" in attribute[0]
                ]
                other_rule_collections = [
                    attribute
                    for attribute in attribute_list
                    # if "application_rule_collection" not in attribute[0]
                    # and "network_rule_collection" not in attribute[0]
                    # and "nat_rule_collection" not in attribute[0]
                ]

                for attribute in other_rule_collections:
                    attribute_path = attribute[0]
                    normalized_path = re.sub(r"\[\d+\]", "", attribute_path)
                    description = descriptions.get(resource_type, {}).get(
                        normalized_path, ""
                    )
                    value = str(attribute[1])
                    ws.append([attribute[0], value, description])
                    if attribute[0] == "priority":
                        collection_group_priority = value

                # application_rule_collectionだけ別のシートに出力
                if application_rule_collections:
                    ws_appcol = wb.create_sheet(title=f"apcols_{collection_group_priority}")
                    # Add header row for collection sheet
                    ws_appcol.append(
                        [
                            "collection_index",
                            "name",
                            "priority",
                            "action",
                        ]
                    )
                    appcol_dict = defaultdict(dict)
                    appcolrule_dict = defaultdict(dict)
                    for app_col in application_rule_collections:
                        app_col_key = re.sub(r"\[\d+\]", "", app_col[0])
                        app_col_index = re.search(r"\[(\d+)\]", app_col[0]).group(1)
                        appcol_dict[app_col_index][app_col_key] = app_col[1]
                        if "rule" in app_col[0]:
                            match = re.search(r"rule\[(\d+)\]", app_col[0])
                            if match:
                                app_col_rule_index = match.group(1)
                                # グループキー：col_index_ruleIndex
                                key = f"{app_col_index}_{app_col_rule_index}"
                                if key in appcolrule_dict:
                                    appcolrule_dict[key][app_col_key] = str(
                                        appcolrule_dict[key].get(app_col_key, "")
                                    ) + f"\n{app_col[1]}"
                                else:
                                    appcolrule_dict[key][app_col_key] = app_col[1]
                    sorted_appcols = sorted(
                        appcol_dict.items(),
                        key=lambda x: int(x[1].get("application_rule_collection.priority", 0)),
                    )
                    for app_col_index, app_col_attrs in sorted_appcols:
                        row = [f"apcol{app_col_index}"]
                        for key in [
                            "application_rule_collection.name",
                            "application_rule_collection.priority",
                            "application_rule_collection.action",
                        ]:
                            row.append(str(app_col_attrs.get(key, "")))
                        ws_appcol.append(row)

                    # apruleシートをcol_indexごとに分ける
                    app_rule_sheets = {}
                    for rule_key, rule_attrs in appcolrule_dict.items():
                        ap_col_index, ap_rule_index = rule_key.split("_", 1)
                        if ap_col_index not in app_rule_sheets:
                            # collectionのpriorityを取得してシート名に含める
                            priority = appcol_dict[ap_col_index].get("application_rule_collection.priority", "0")
                            app_rule_sheets[ap_col_index] = wb.create_sheet(title=f"apcol_{collection_group_priority}_{priority}_rules")
                            app_rule_sheets[ap_col_index].append(
                                [
                                    "rule_Index",
                                    "name",
                                    "description",
                                    "destination_addresses",
                                    "destination_fqdn_tags",
                                    "destination_fqdns",
                                    "destination_urls",
                                    "http_headers",
                                    "protocols",
                                    "source_addresses",
                                    "source_ip_groups",
                                    "terminate_tls",
                                    "web_categories",
                                ]
                            )
                        row = [f"aprule{ap_rule_index}".strip()]
                        for key in [
                            "application_rule_collection.rule.name",
                            "application_rule_collection.rule.description",
                            "application_rule_collection.rule.destination_addresses",
                            "application_rule_collection.rule.destination_fqdn_tags",
                            "application_rule_collection.rule.destination_fqdns",
                            "application_rule_collection.rule.destination_urls",
                            "application_rule_collection.rule.http_headers",
                            "application_rule_collection.rule.protocols",
                            "application_rule_collection.rule.source_addresses",
                            "application_rule_collection.rule.source_ip_groups",
                            "application_rule_collection.rule.terminate_tls",
                            "application_rule_collection.rule.web_categories",
                        ]:
                            if key == "application_rule_collection.rule.protocols":
                                ports = str(rule_attrs.get("application_rule_collection.rule.protocols.port", "")).split("\n")
                                types = str(rule_attrs.get("application_rule_collection.rule.protocols.type", "")).split("\n")
                                protocols = [f"{t}:{p}" for p, t in zip(ports, types) if p.strip() or t.strip()]
                                row.append("\n".join(protocols))
                            else:
                                row.append(str(rule_attrs.get(key, "")).strip())
                        app_rule_sheets[ap_col_index].append(row)

                    apply_styles(
                        ws_appcol,
                        header_fill,
                        header_font,
                        default_font,
                        left_alignment,
                        thin_border,
                    )
                    for sheet in app_rule_sheets.values():
                        apply_styles(
                            sheet,
                            header_fill,
                            header_font,
                            default_font,
                            left_alignment,
                            thin_border,
                        )
                    # -------------------------------------------------------

                # network_rule_collectionだけ別のシートに出力
                if network_rule_collections:
                    ws_netcol = wb.create_sheet(title=f"netcols_{collection_group_priority}")
                    # Add header row for network_rule_collection sheet
                    ws_netcol.append(
                        [
                            "collection_index",
                            "name",
                            "priority",
                            "action",
                        ]
                    )
                    netcol_dict = defaultdict(dict)
                    netcolrule_dict = defaultdict(dict)
                    for net_col in network_rule_collections:
                        net_col_key = re.sub(r"\[\d+\]", "", net_col[0])
                        net_col_index = re.search(r"\[(\d+)\]", net_col[0]).group(1)
                        netcol_dict[net_col_index][net_col_key] = net_col[1]
                        if "rule" in net_col[0]:
                            match = re.search(r"rule\[(\d+)\]", net_col[0])
                            if match:
                                net_col_rule_index = match.group(1)
                                key = f"{net_col_index}_{net_col_rule_index}"
                                if key in netcolrule_dict:
                                    netcolrule_dict[key][net_col_key] = str(netcolrule_dict[key].get(net_col_key, "")) + f"\n{net_col[1]}"
                                else:
                                    netcolrule_dict[key][net_col_key] = net_col[1]
                    # Sort net_cols by priority
                    sorted_netcols = sorted(
                        netcol_dict.items(),
                        key=lambda x: int(x[1].get("network_rule_collection.priority", 0)),
                    )
                    for net_col_index, net_col_attrs in sorted_netcols:
                        row = [f"netcol{net_col_index}"]
                        for key in [
                            "network_rule_collection.name",
                            "network_rule_collection.priority",
                            "network_rule_collection.action",
                        ]:
                            row.append(str(net_col_attrs.get(key, "")))
                        ws_netcol.append(row)
                    
                    # ----- 変更箇所: net_ruleシートを個別に作成する -----
                    net_rule_sheets = {}
                    for net_key, net_attrs in netcolrule_dict.items():
                        net_col_index, net_rule_index = net_key.split("_", 1)
                        if net_col_index not in net_rule_sheets:
                            priority = netcol_dict[net_col_index].get("network_rule_collection.priority", "0")
                            net_rule_sheets[net_col_index] = wb.create_sheet(title=f"netcol_{collection_group_priority}_{priority}_rules")
                            net_rule_sheets[net_col_index].append(
                                [
                                    "rule_Index",
                                    "name",
                                    "description",
                                    "destination_addresses",
                                    "destination_fqdns",
                                    "destination_ip_groups",
                                    "destination_ports",
                                    "protocols",
                                    "source_addresses",
                                    "source_ip_groups",
                                ]
                            )
                        row = [f"netrule{net_rule_index}".strip()]
                        for key in [
                            "network_rule_collection.rule.name",
                            "network_rule_collection.rule.description",
                            "network_rule_collection.rule.destination_addresses",
                            "network_rule_collection.rule.destination_fqdns",
                            "network_rule_collection.rule.destination_ip_groups",
                            "network_rule_collection.rule.destination_ports",
                            "network_rule_collection.rule.protocols",
                            "network_rule_collection.rule.source_addresses",
                            "network_rule_collection.rule.source_ip_groups",
                        ]:
                            row.append(str(net_attrs.get(key, "")).strip())
                        net_rule_sheets[net_col_index].append(row)
                    
                    apply_styles(
                        ws_netcol,
                        header_fill,
                        header_font,
                        default_font,
                        left_alignment,
                        thin_border,
                    )
                    for sheet in net_rule_sheets.values():
                        apply_styles(
                            sheet,
                            header_fill,
                            header_font,
                            default_font,
                            left_alignment,
                            thin_border,
                        )
                    # -------------------------------------------------------

                # nat_rule_collectionだけ別のシートに出力
                if nat_rule_collections:
                    ws_natcol = wb.create_sheet(title=f"natcols_{collection_group_priority}")
                    # Add header row for nat_rule_collection sheet
                    ws_natcol.append(
                        [
                            "collection_index",
                            "name",
                            "priority",
                            "action",
                        ]
                    )
                    natcol_dict = defaultdict(dict)
                    natcolrule_dict = defaultdict(dict)
                    for nat_col in nat_rule_collections:
                        nat_col_key = re.sub(r"\[\d+\]", "", nat_col[0])
                        nat_col_index = re.search(r"\[(\d+)\]", nat_col[0]).group(1)
                        natcol_dict[nat_col_index][nat_col_key] = nat_col[1]
                        if "rule" in nat_col[0]:
                            match = re.search(r"rule\[(\d+)\]", nat_col[0])
                            if match:
                                nat_col_rule_index = match.group(1)
                                key = f"{nat_col_index}_{nat_col_rule_index}"
                                if key in natcolrule_dict:
                                    natcolrule_dict[key][nat_col_key] = str(natcolrule_dict[key].get(nat_col_key, "")) + f"\n{nat_col[1]}"
                                else:
                                    natcolrule_dict[key][nat_col_key] = nat_col[1]
                    # Sort nat_cols by priority
                    sorted_natcols = sorted(
                        natcol_dict.items(),
                        key=lambda x: int(x[1].get("nat_rule_collection.priority", 0)),
                    )
                    for nat_col_index, nat_col_attrs in sorted_natcols:
                        row = [f"natcol{nat_col_index}"]
                        for key in [
                            "nat_rule_collection.name",
                            "nat_rule_collection.priority",
                            "nat_rule_collection.action",
                        ]:
                            row.append(str(nat_col_attrs.get(key, "")))
                        ws_natcol.append(row)
                    
                    # ----- 変更箇所: nat_ruleシートを個別に作成する -----
                    nat_rule_sheets = {}
                    for nat_key, nat_attrs in natcolrule_dict.items():
                        nat_col_index, nat_rule_index = nat_key.split("_", 1)
                        if nat_col_index not in nat_rule_sheets:
                            priority = natcol_dict[nat_col_index].get("nat_rule_collection.priority", "0")
                            nat_rule_sheets[nat_col_index] = wb.create_sheet(title=f"natcol_{collection_group_priority}_{priority}_rules")
                            nat_rule_sheets[nat_col_index].append(
                                [
                                    "rule_index",
                                    "name",
                                    "description",
                                    "destination_address",
                                    "destination_ports",
                                    "protocols",
                                    "source_addresses",
                                    "source_ip_groups",
                                    "translated_address",
                                    "translated_fqdn",
                                    "translated_port",
                                ]
                            )
                        row = [f"natrule{nat_rule_index}".strip()]
                        for key in [
                            "nat_rule_collection.rule.name",
                            "nat_rule_collection.rule.description",
                            "nat_rule_collection.rule.destination_address",
                            "nat_rule_collection.rule.destination_ports",
                            "nat_rule_collection.rule.protocols",
                            "nat_rule_collection.rule.source_addresses",
                            "nat_rule_collection.rule.source_ip_groups",
                            "nat_rule_collection.rule.translated_address",
                            "nat_rule_collection.rule.translated_fqdn",
                            "nat_rule_collection.rule.translated_port",
                        ]:
                            row.append(str(nat_attrs.get(key, "")).strip())
                        nat_rule_sheets[nat_col_index].append(row)
                    apply_styles(
                        ws_natcol,
                        header_fill,
                        header_font,
                        default_font,
                        left_alignment,
                        thin_border,
                    )
                    for sheet in nat_rule_sheets.values():
                        apply_styles(
                            sheet,
                            header_fill,
                            header_font,
                            default_font,
                            left_alignment,
                            thin_border,
                        )
                    # -------------------------------------------------------

            else:
                for attribute in attribute_list:
                    attribute_path = attribute[0]
                    normalized_path = re.sub(r"\[\d+\]", "", attribute_path)
                    description = descriptions.get(resource_type, {}).get(
                        normalized_path, ""
                    )
                    value = str(attribute[1])
                    ws.append([attribute[0], value, description])

            # Apply font and alignment to all other cells
            default_font = Font(name="Yu Gothic UI")
            left_alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )  # Left alignment with text wrapping and top alignment
            for row in ws.iter_rows(min_row=2):  # Skip header row
                for cell in row:
                    cell.font = default_font
                    cell.alignment = left_alignment

            # Apply border to all cells
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border

            # Adjust column widths
            ws.column_dimensions["B"].width = 100  # Set width for Value column
            ws.column_dimensions["C"].width = 100  # Set width for Description column

            # Adjust column widths for non-B,C columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                if column not in ["B", "C"]:  # Skip Value and Description columns
                    for cell in col:
                        try:
                            if cell.value is not None:
                                lines = str(cell.value).splitlines()
                                cell_max = max(len(line) for line in lines) if lines else 0
                                if cell_max > max_length:
                                    max_length = cell_max
                        except:
                            pass
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column].width = adjusted_width

        # Remove the default sheet created by openpyxl
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        output_path = os.path.join(output_folder, f"{resource_type}.xlsx")
        wb.save(output_path)
        print(f"Excel file saved to {output_path}")


def process_tfstate(tfstate_file, description_folders, output_folder):
    with open(tfstate_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    resources_by_type = defaultdict(list)
    for res in data["resources"]:
        if res["mode"] == "managed":
            resources_by_type[res["type"]].append(res)

    descriptions = load_descriptions(description_folders)
    write_to_excel(resources_by_type, descriptions, output_folder)


if __name__ == "__main__":
    if len(sys.argv) < 3:  # 引数の数を変更
        print(
            "Usage: python azurerm2excel.py <path_to_tfstate_file> <description_folder1> [<description_folder2> ...]"
        )  # メッセージを変更
        sys.exit(1)

    tfstate_file = sys.argv[1]
    description_folders = sys.argv[
        2:
    ]  # description_folder を description_folders に変更
    output_folder = "output"  # 固定の出力フォルダ

    if not os.path.isfile(tfstate_file):
        print(f"Error: The file {tfstate_file} does not exist。")
        sys.exit(1)

    for description_folder in description_folders:  # 各フォルダをチェック
        if not os.path.isdir(description_folder):
            print(f"Error: The directory {description_folder} does not exist。")
            sys.exit(1)

    # フォルダが存在しない場合は作成する
    if not os.path.isdir(output_folder):
        print(f"Creating output folder: {output_folder}")
        os.makedirs(output_folder)

    # YYYYMMDDss フォルダを作成
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    output_folder = os.path.join(output_folder, timestamp)
    os.makedirs(output_folder)

    process_tfstate(
        tfstate_file, description_folders, output_folder
    )
